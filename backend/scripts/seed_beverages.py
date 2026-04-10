"""Seed beverages from Excel file into the database as global entries."""
import asyncio
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy import select, text
from app.core.database import async_session_factory
from app.models.beverage import Beverage


def parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.replace(",", "."))
        except ValueError:
            return 0.0
    return 0.0


async def seed():
    wb = openpyxl.load_workbook(
        os.path.join(os.path.dirname(__file__), "..", "..", "documentation", "Excel Maestro de Bebidas Tracfiz.xlsx")
    )
    ws = wb.active

    beverages = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name = row[0].value
        if not name:
            continue
        beverages.append({
            "name": str(name).strip(),
            "category": str(row[1].value).strip() if row[1].value else None,
            "serving_size_ml": parse_number(row[2].value),
            "reference_ml": parse_number(row[3].value),
            "calories": parse_number(row[4].value),
            "protein": parse_number(row[5].value),
            "fat": parse_number(row[6].value),
            "carbs": parse_number(row[7].value),
        })

    async with async_session_factory() as session:
        existing = await session.execute(
            select(Beverage).where(Beverage.is_global.is_(True)).limit(1)
        )
        if existing.scalar_one_or_none():
            print("Global beverages already exist, skipping seed.")
            return

        for b in beverages:
            session.add(Beverage(
                name=b["name"],
                category=b["category"],
                serving_size_ml=b["serving_size_ml"],
                reference_ml=b["reference_ml"],
                calories=b["calories"],
                protein=b["protein"],
                fat=b["fat"],
                carbs=b["carbs"],
                is_global=True,
                workspace_id=None,
            ))

        await session.commit()
        print(f"Seeded {len(beverages)} global beverages.")


if __name__ == "__main__":
    asyncio.run(seed())
