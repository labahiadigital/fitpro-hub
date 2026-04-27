"""Stock management API endpoints."""
import io
import logging
from datetime import date
from decimal import Decimal
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel as PydanticModel
from sqlalchemy import select, func, and_, cast, Date
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.middleware.auth import require_workspace
from app.models.stock import StockCategory, StockItem, StockItemBox, StockMovement
from app.models.product import ProductStockConsumption, Product
from app.models.resource import ServiceStockConsumption, Service
from app.models.resource import Box

logger = logging.getLogger(__name__)
router = APIRouter()

CurrentUser = Depends(require_workspace)


_DEFERRED_STOCK_ATTRS = {"box_id", "supplier_id"}


def _safe_attr(obj, name: str):
    """Lee un atributo defensivamente SIN disparar lazy-load en deferred.

    Si la columna no existe aún en el schema (p.ej. migración 048 pendiente),
    ``getattr`` sobre un atributo ``deferred`` lanzaría un SELECT contra esa
    columna y explotaría con ``UndefinedColumn`` abortando la transacción.
    Para los atributos conocidos como diferidos leemos directamente de
    ``__dict__`` (que sólo contiene valores ya cargados por SQLAlchemy).
    """
    if name in _DEFERRED_STOCK_ATTRS:
        return getattr(obj, "__dict__", {}).get(name)
    try:
        return getattr(obj, name, None)
    except Exception:
        return None


def _safe_set(obj, name: str, value) -> None:
    try:
        setattr(obj, name, value)
    except Exception:
        pass


async def _load_box_allocations_map(
    db: AsyncSession, workspace_id, item_ids: List[UUID]
) -> dict:
    """Devuelve un dict ``{item_id: [BoxAllocationOut...]}`` para los items dados.

    Si la tabla ``stock_item_boxes`` aún no existe (migración 049 no aplicada),
    devuelve un dict vacío para no romper el listado.
    """
    if not item_ids:
        return {}
    try:
        result = await db.execute(
            select(StockItemBox, Box.name)
            .join(Box, Box.id == StockItemBox.box_id, isouter=True)
            .where(
                StockItemBox.workspace_id == workspace_id,
                StockItemBox.item_id.in_(item_ids),
            )
        )
    except Exception as exc:  # pragma: no cover - migración pendiente
        logger.warning("box_allocations: tabla aún no disponible (%s)", exc)
        try:
            await db.rollback()
        except Exception:
            pass
        return {}
    bucket: dict = {}
    for alloc, box_name in result.all():
        bucket.setdefault(alloc.item_id, []).append({
            "box_id": alloc.box_id,
            "box_name": box_name,
            "current_stock": float(alloc.current_stock),
            "min_stock": float(alloc.min_stock),
            "max_stock": float(alloc.max_stock),
        })
    return bucket


async def _sync_box_allocations(
    db: AsyncSession,
    workspace_id,
    item: StockItem,
    allocations: List[BoxAllocationInput],
) -> Optional[float]:
    """Reemplaza las allocations del item con las recibidas.

    Devuelve la suma de unidades agregadas para que el llamador pueda
    actualizar ``StockItem.current_stock`` en consecuencia. Si la tabla aún
    no existe (migración pendiente), devuelve ``None`` para que el caller
    decida no tocar el current_stock derivado.
    """
    try:
        existing = await db.execute(
            select(StockItemBox).where(StockItemBox.item_id == item.id)
        )
        for row in existing.scalars().all():
            await db.delete(row)
        await db.flush()
    except Exception as exc:  # pragma: no cover - migración pendiente
        logger.warning("sync_box_allocations: tabla no disponible (%s)", exc)
        try:
            await db.rollback()
        except Exception:
            pass
        return None

    seen: set = set()
    total = 0.0
    for alloc in allocations or []:
        if alloc.box_id in seen:
            continue
        seen.add(alloc.box_id)
        row = StockItemBox(
            workspace_id=workspace_id,
            item_id=item.id,
            box_id=alloc.box_id,
            current_stock=Decimal(str(alloc.current_stock or 0)),
            min_stock=Decimal(str(alloc.min_stock or 0)),
            max_stock=Decimal(str(alloc.max_stock or 0)),
        )
        db.add(row)
        total += float(alloc.current_stock or 0)
    return total


# --- Pydantic schemas ---

class CategoryCreate(PydanticModel):
    name: str
    icon: Optional[str] = None

class CategoryResponse(PydanticModel):
    id: UUID
    name: str
    icon: Optional[str] = None

    class Config:
        from_attributes = True

class BoxAllocationInput(PydanticModel):
    box_id: UUID
    current_stock: float = 0
    min_stock: float = 0
    max_stock: float = 0


class BoxAllocationOut(PydanticModel):
    box_id: UUID
    box_name: Optional[str] = None
    current_stock: float
    min_stock: float
    max_stock: float


class ItemCreate(PydanticModel):
    name: str
    category_id: Optional[UUID] = None
    description: Optional[str] = None
    unit: str = "ud"
    current_stock: float = 0
    min_stock: float = 0
    max_stock: float = 0
    price: float = 0
    location: Optional[str] = None
    tax_rate: float = 21
    irpf_rate: float = 0
    box_id: Optional[UUID] = None
    supplier_id: Optional[UUID] = None
    box_allocations: Optional[List[BoxAllocationInput]] = None

class ItemUpdate(PydanticModel):
    name: Optional[str] = None
    category_id: Optional[UUID] = None
    description: Optional[str] = None
    unit: Optional[str] = None
    min_stock: Optional[float] = None
    max_stock: Optional[float] = None
    price: Optional[float] = None
    location: Optional[str] = None
    tax_rate: Optional[float] = None
    irpf_rate: Optional[float] = None
    box_id: Optional[UUID] = None
    supplier_id: Optional[UUID] = None
    box_allocations: Optional[List[BoxAllocationInput]] = None

class MovementCreate(PydanticModel):
    movement_type: str  # entry, exit, adjustment
    quantity: float
    reason: str
    box_id: Optional[UUID] = None

class ItemResponse(PydanticModel):
    id: UUID
    name: str
    category_id: Optional[UUID] = None
    category_name: Optional[str] = None
    description: Optional[str] = None
    unit: str
    current_stock: float
    min_stock: float
    max_stock: float
    price: float
    location: Optional[str] = None
    tax_rate: float
    irpf_rate: float
    is_active: bool
    created_at: str

class MovementResponse(PydanticModel):
    id: UUID
    item_id: UUID
    movement_type: str
    quantity: float
    previous_stock: float
    new_stock: float
    reason: str
    created_by: Optional[UUID] = None
    created_at: str

class StockSummary(PydanticModel):
    total_items: int
    low_stock_count: int
    total_value: float
    movements_today: int


# --- Categories ---

@router.get("/categories", response_model=List[CategoryResponse])
async def list_categories(user=CurrentUser, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockCategory)
        .where(StockCategory.workspace_id == user.workspace_id)
        .order_by(StockCategory.name)
    )
    return result.scalars().all()


@router.post("/categories", response_model=CategoryResponse)
async def create_category(data: CategoryCreate, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    cat = StockCategory(workspace_id=user.workspace_id, name=data.name, icon=data.icon)
    db.add(cat)
    await db.commit()
    await db.refresh(cat)
    return cat


# --- Items ---

@router.get("/items")
async def list_items(
    search: str = "",
    category_id: Optional[str] = None,
    low_stock: bool = False,
    user=CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    q = (
        select(StockItem)
        .options(selectinload(StockItem.category))
        .where(StockItem.workspace_id == user.workspace_id, StockItem.is_active.is_(True))
    )
    if search:
        q = q.where(StockItem.name.ilike(f"%{search}%"))
    if category_id:
        q = q.where(StockItem.category_id == category_id)
    if low_stock:
        q = q.where(StockItem.current_stock <= StockItem.min_stock)
    q = q.order_by(StockItem.name)
    result = await db.execute(q)
    items = result.scalars().all()

    allocations_by_item = await _load_box_allocations_map(
        db, user.workspace_id, [i.id for i in items]
    )

    def _serialize_alloc(alloc: dict) -> dict:
        return {
            "box_id": str(alloc["box_id"]),
            "box_name": alloc.get("box_name"),
            "current_stock": alloc["current_stock"],
            "min_stock": alloc["min_stock"],
            "max_stock": alloc["max_stock"],
        }

    out: List[dict] = []
    for i in items:
        boxes = [_serialize_alloc(a) for a in allocations_by_item.get(i.id, [])]
        # Si tiene multi-box, el stock total se considera la suma de boxes;
        # si no hay allocations, se usa el current_stock directo del item.
        derived_stock = sum(b["current_stock"] for b in boxes) if boxes else float(i.current_stock)
        out.append({
            "id": str(i.id),
            "name": i.name,
            "category_id": str(i.category_id) if i.category_id else None,
            "category_name": i.category.name if i.category else None,
            "description": i.description,
            "unit": i.unit,
            "current_stock": derived_stock,
            "min_stock": float(i.min_stock),
            "max_stock": float(i.max_stock),
            "price": float(i.price),
            "location": i.location,
            "tax_rate": float(i.tax_rate),
            "irpf_rate": float(i.irpf_rate),
            "is_active": i.is_active,
            "box_id": str(_safe_attr(i, "box_id")) if _safe_attr(i, "box_id") else None,
            "supplier_id": str(_safe_attr(i, "supplier_id")) if _safe_attr(i, "supplier_id") else None,
            "box_allocations": boxes,
            "created_at": i.created_at.isoformat() if i.created_at else "",
        })
    return out


def _build_stock_item_base(data: "ItemCreate", workspace_id) -> StockItem:
    return StockItem(
        workspace_id=workspace_id,
        name=data.name,
        category_id=data.category_id,
        description=data.description,
        unit=data.unit,
        current_stock=data.current_stock,
        min_stock=data.min_stock,
        max_stock=data.max_stock,
        price=data.price,
        location=data.location,
        tax_rate=data.tax_rate,
        irpf_rate=data.irpf_rate,
    )


@router.post("/items")
async def create_item(data: ItemCreate, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    item = _build_stock_item_base(data, user.workspace_id)
    # box/supplier son opcionales y pueden no existir aún como columnas en
    # entornos donde no se haya aplicado la migración 048.
    if data.box_id is not None:
        _safe_set(item, "box_id", data.box_id)
    if data.supplier_id is not None:
        _safe_set(item, "supplier_id", data.supplier_id)
    db.add(item)
    try:
        await db.commit()
    except Exception as exc:  # noqa: BLE001
        try:
            await db.rollback()
        except Exception:  # noqa: BLE001
            pass
        logger.warning("create_item: retry sin box/supplier (%s)", exc)
        retry = _build_stock_item_base(data, user.workspace_id)
        db.add(retry)
        await db.commit()
        item = retry
    await db.refresh(item)

    if data.box_allocations:
        total = await _sync_box_allocations(db, user.workspace_id, item, data.box_allocations)
        if total is not None:
            item.current_stock = Decimal(str(total))
        await db.commit()
        await db.refresh(item)
    return {"id": str(item.id), "name": item.name}


@router.put("/items/{item_id}")
async def update_item(item_id: UUID, data: ItemUpdate, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    payload = data.model_dump(exclude_unset=True)
    box_allocations = payload.pop("box_allocations", None)
    touched_deferred = False
    if "box_id" in payload:
        _safe_set(item, "box_id", payload.pop("box_id"))
        touched_deferred = True
    if "supplier_id" in payload:
        _safe_set(item, "supplier_id", payload.pop("supplier_id"))
        touched_deferred = True
    for field, value in payload.items():
        setattr(item, field, value)
    try:
        await db.commit()
    except Exception as exc:  # noqa: BLE001
        try:
            await db.rollback()
        except Exception:  # noqa: BLE001
            pass
        if not touched_deferred:
            raise
        logger.warning("update_item: retry sin box/supplier (%s)", exc)
        result = await db.execute(
            select(StockItem).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise HTTPException(404, "Item not found")
        for field, value in data.model_dump(exclude_unset=True).items():
            if field in _DEFERRED_STOCK_ATTRS or field == "box_allocations":
                continue
            setattr(item, field, value)
        await db.commit()

    if box_allocations is not None:
        allocations_input = [BoxAllocationInput(**a) if isinstance(a, dict) else a for a in box_allocations]
        total = await _sync_box_allocations(db, user.workspace_id, item, allocations_input)
        if total is not None:
            item.current_stock = Decimal(str(total))
        await db.commit()
    return {"ok": True}


@router.delete("/items/{item_id}")
async def delete_item(item_id: UUID, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    item.is_active = False
    await db.commit()
    return {"ok": True}


# --- Movements ---

@router.post("/items/{item_id}/movements")
async def register_movement(item_id: UUID, data: MovementCreate, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    qty = data.quantity
    if data.movement_type not in {"entry", "exit", "adjustment"}:
        raise HTTPException(400, "Invalid movement type")

    # Si el movimiento se aplica a un box concreto, ajustamos esa allocation
    # y derivamos el current_stock global como suma. Si no se especifica
    # box, aplicamos al item directamente (modo legacy).
    box_alloc: Optional[StockItemBox] = None
    if data.box_id is not None:
        try:
            res_alloc = await db.execute(
                select(StockItemBox).where(
                    StockItemBox.item_id == item.id,
                    StockItemBox.box_id == data.box_id,
                )
            )
            box_alloc = res_alloc.scalar_one_or_none()
        except Exception:  # tabla no creada aún
            box_alloc = None

    if box_alloc is not None:
        previous = float(box_alloc.current_stock)
    else:
        previous = float(item.current_stock)

    if data.movement_type == "entry":
        new_stock = previous + qty
    elif data.movement_type == "exit":
        new_stock = max(0, previous - qty)
    else:  # adjustment
        new_stock = qty

    if box_alloc is not None:
        box_alloc.current_stock = Decimal(str(new_stock))
        await db.flush()
        # Recalcula el current_stock total del item a partir de allocations.
        total_q = await db.execute(
            select(func.coalesce(func.sum(StockItemBox.current_stock), 0))
            .where(StockItemBox.item_id == item.id)
        )
        total = float(total_q.scalar() or 0)
        item.current_stock = Decimal(str(total))
    else:
        item.current_stock = Decimal(str(new_stock))

    movement = StockMovement(
        workspace_id=user.workspace_id,
        item_id=item_id,
        movement_type=data.movement_type,
        quantity=Decimal(str(qty)),
        previous_stock=Decimal(str(previous)),
        new_stock=Decimal(str(new_stock)),
        reason=data.reason,
        created_by=user.id,
    )
    if data.box_id is not None:
        _safe_set(movement, "box_id", data.box_id)
    db.add(movement)
    await db.commit()
    return {"ok": True, "new_stock": float(item.current_stock)}


# --- Box allocations ---


@router.get("/items/{item_id}/boxes", response_model=List[BoxAllocationOut])
async def list_box_allocations(item_id: UUID, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    item_check = await db.execute(
        select(StockItem.id).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
    )
    if not item_check.scalar_one_or_none():
        raise HTTPException(404, "Item not found")
    allocations = await _load_box_allocations_map(db, user.workspace_id, [item_id])
    return [
        BoxAllocationOut(
            box_id=a["box_id"],
            box_name=a.get("box_name"),
            current_stock=a["current_stock"],
            min_stock=a["min_stock"],
            max_stock=a["max_stock"],
        )
        for a in allocations.get(item_id, [])
    ]


@router.put("/items/{item_id}/boxes")
async def upsert_box_allocations(
    item_id: UUID,
    data: List[BoxAllocationInput],
    user=CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    """Reemplaza la distribución por boxes del item con las allocations dadas.

    El ``current_stock`` total del item pasa a ser la suma de las unidades
    asignadas. Si el listado llega vacío, se borran todas las allocations y
    el item vuelve al modo legacy (un único almacén implícito).
    """
    res = await db.execute(
        select(StockItem).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
    )
    item = res.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    total = await _sync_box_allocations(db, user.workspace_id, item, data)
    if total is None:
        raise HTTPException(503, "La distribución por boxes aún no está disponible. Aplica la migración 049.")
    item.current_stock = Decimal(str(total))
    await db.commit()
    return {"ok": True, "current_stock": total, "boxes": len(data)}


@router.get("/items/{item_id}/movements")
async def list_movements(item_id: UUID, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(StockMovement)
        .where(StockMovement.item_id == item_id, StockMovement.workspace_id == user.workspace_id)
        .order_by(StockMovement.created_at.desc())
        .limit(100)
    )
    movements = result.scalars().all()
    return [
        {
            "id": str(m.id),
            "item_id": str(m.item_id),
            "movement_type": m.movement_type,
            "quantity": float(m.quantity),
            "previous_stock": float(m.previous_stock),
            "new_stock": float(m.new_stock),
            "reason": m.reason,
            "created_by": str(m.created_by) if m.created_by else None,
            "created_at": m.created_at.isoformat() if m.created_at else "",
        }
        for m in movements
    ]


# --- Summary / KPIs ---

@router.get("/summary")
async def get_summary(user=CurrentUser, db: AsyncSession = Depends(get_db)):
    ws = user.workspace_id

    items_result = await db.execute(
        select(
            func.count(StockItem.id).label("total"),
            func.count().filter(StockItem.current_stock <= StockItem.min_stock).label("low"),
            func.coalesce(func.sum(StockItem.current_stock * StockItem.price), 0).label("value"),
        ).where(StockItem.workspace_id == ws, StockItem.is_active.is_(True))
    )
    row = items_result.one()

    today = date.today()
    mov_result = await db.execute(
        select(func.count(StockMovement.id)).where(
            StockMovement.workspace_id == ws,
            cast(StockMovement.created_at, Date) == today,
        )
    )
    movements_today = mov_result.scalar() or 0

    return {
        "total_items": row.total or 0,
        "low_stock_count": row.low or 0,
        "total_value": float(row.value or 0),
        "movements_today": movements_today,
    }


# --- Linked Products for a stock item ---

@router.get("/items/{item_id}/linked-products")
async def get_linked_products(item_id: UUID, user=CurrentUser, db: AsyncSession = Depends(get_db)):
    """Get products and services linked to a stock item via consumption tables."""
    item_check = await db.execute(
        select(StockItem.id).where(StockItem.id == item_id, StockItem.workspace_id == user.workspace_id)
    )
    if not item_check.scalar_one_or_none():
        raise HTTPException(404, "Item not found")

    product_links = await db.execute(
        select(ProductStockConsumption, Product.name)
        .join(Product, Product.id == ProductStockConsumption.product_id)
        .where(
            ProductStockConsumption.stock_item_id == item_id,
            Product.workspace_id == user.workspace_id,
        )
    )
    products = [
        {
            "type": "product",
            "id": str(row[0].product_id),
            "name": row[1],
            "quantity_per_sale": float(row[0].quantity),
        }
        for row in product_links
    ]

    service_links = await db.execute(
        select(ServiceStockConsumption, Service.name)
        .join(Service, Service.id == ServiceStockConsumption.service_id)
        .where(
            ServiceStockConsumption.stock_item_id == item_id,
            Service.workspace_id == user.workspace_id,
        )
    )
    services = [
        {
            "type": "service",
            "id": str(row[0].service_id),
            "name": row[1],
            "quantity_per_sale": float(row[0].quantity),
        }
        for row in service_links
    ]

    return products + services


# --- Export stock to Excel ---

@router.get("/export")
async def export_stock_excel(user=CurrentUser, db: AsyncSession = Depends(get_db)):
    """Export all stock items to an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    q = (
        select(StockItem)
        .options(selectinload(StockItem.category))
        .where(StockItem.workspace_id == user.workspace_id, StockItem.is_active.is_(True))
        .order_by(StockItem.name)
    )
    result = await db.execute(q)
    items = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["Nombre", "Categoría", "Descripción", "Unidad", "Stock Actual", "Stock Mín.", "Stock Máx.", "Precio (€)", "Valor Total (€)", "Ubicación", "IVA %", "IRPF %", "Estado"]
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        stock = float(item.current_stock)
        price = float(item.price)
        values = [
            item.name,
            item.category.name if item.category else "",
            item.description or "",
            item.unit,
            stock,
            float(item.min_stock),
            float(item.max_stock),
            price,
            round(stock * price, 2),
            item.location or "",
            float(item.tax_rate),
            float(item.irpf_rate),
            "Bajo" if stock <= float(item.min_stock) else "Normal",
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=stock_{date.today().isoformat()}.xlsx"},
    )
